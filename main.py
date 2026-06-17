import io
from fastapi import FastAPI, Query
from fastapi.responses import StreamingResponse
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

app = FastAPI(
    title="Automated Bookstore Scraper API",
    description=(
        "A FastAPI-driven web service that scrapes listings from an online bookstore and "
        "returns structured book data. Change the page parameter to pull listings from "
        "different sections of the catalog and view updated results in real time."
    )
)

# HELPER FUNCTION: Reusable scraping logic
def scrape_books_data(pages: int):
    books_list = []
    for page in range(1, pages + 1):
        url = f"https://books.toscrape.com/catalogue/page-{page}.html"
        response = requests.get(url)
        if response.status_code != 200:
            continue
            
        soup = BeautifulSoup(response.text, "html.parser")
        articles = soup.find_all("article", class_="product_pod")
        
        for article in articles:
            title = article.h3.a["title"]
            price_text = article.find("p", class_="price_color").text
            # Clean and isolate numeric value (e.g., £51.77 -> 51.77)
            clean_price = price_text.replace("Â", "").replace("£", "").strip()
            
            books_list.append({
                "title": title,
                "price": float(clean_price)
            })
    return books_list

# ENDPOINT 1: Returns raw JSON
@app.get("/api/v1/books")
def get_scraped_books(pages: int = Query(default=1, ge=1, le=10, description="The number of pages to scrape")):
    books = scrape_books_data(pages)
    formatted_books = [{"title": b["title"], "price": f"£{b['price']:.2f}"} for b in books]
    return {
        "status": "success", 
        "pages_scraped": pages,
        "count": len(books), 
        "data": formatted_books
    }

# ENDPOINT 2: Generates and streams the dynamic Excel dashboard file!
@app.get("/api/v1/books/download")
def download_excel_dashboard(pages: int = Query(default=1, ge=1, le=10, description="Select scraping depth for the Excel file")):
    # 1. Fetch live data based on recruiter input
    books = scrape_books_data(pages)
    
    # 2. Build the Excel spreadsheet in code
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookstore Dashboard"
    
    # Enable gridlines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Headers
    headers = ["Book Title", "Price (GBP)"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_num == 2 else "left")
    
    # Append Data Rows
    for i, book in enumerate(books, 2):
        ws.append([book["title"], book["price"]])
        
        # Style formatting
        cell_title = ws.cell(row=i, column=1)
        cell_price = ws.cell(row=i, column=2)
        
        cell_title.font = data_font
        cell_price.font = data_font
        cell_price.number_format = '"£"#,##0.00'
        cell_price.alignment = Alignment(horizontal="right")
        
        # Zebra striping on alternating rows
        if i % 2 == 1:
            cell_title.fill = zebra_fill
            cell_price.fill = zebra_fill

    # 🛠️ FIXED: Dynamic calculation of data bounds
    total_data_rows = len(books)
    actual_last_data_row = total_data_rows + 1 # +1 accounts for the header row
    
    summary_start = actual_last_data_row + 2 # Leave a blank row spacer
    
    # Write Total Items Formula
    ws.cell(row=summary_start, column=1, value="Total Unique Items Logged").font = Font(name="Segoe UI", bold=True)
    count_cell = ws.cell(row=summary_start, column=2, value=f"=COUNTA(B2:B{actual_last_data_row})")
    count_cell.font = Font(name="Segoe UI", bold=True)
    count_cell.alignment = Alignment(horizontal="right")
    
    # Write Average Value Formula
    ws.cell(row=summary_start+1, column=1, value="Average Book Value").font = Font(name="Segoe UI", bold=True)
    avg_cell = ws.cell(row=summary_start+1, column=2, value=f"=AVERAGE(B2:B{actual_last_data_row})")
    avg_cell.font = Font(name="Segoe UI", bold=True)
    avg_cell.number_format = '"£"#,##0.00'
    avg_cell.alignment = Alignment(horizontal="right")

    # 📊 ADDED: Programmatic Generation of the Visual Bar Chart
    # 📊 UPDATED: Programmatic Generation of the Visual Bar Chart with Explicit Title Rendering
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    
    # Force Excel to explicitly render the title text layout
    chart.title = "Book Price Distribution Landscape"
    
    chart.y_axis.title = "Price (GBP)"
    chart.x_axis.title = "Books"
    
    # Tell the chart to reference the numerical price data column
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=actual_last_data_row)
    # Tell the chart to use titles from column 1 as data labels
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=actual_last_data_row)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    # Explicit layout settings to prevent title stripping
    chart.legend = None 
    chart.width = 18   # Making it slightly wider so names fit better
    chart.height = 10  # Giving it clean vertical room
    
    # Position chart on the right-hand side so it doesn't cover data rows
    chart_placement_cell = "D2"
    ws.add_chart(chart, chart_placement_cell)

    # Auto-adjust column widths based on longest string lengths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 3. Save workbook directly to a RAM memory stream instead of writing to disk
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # 4. Stream the file back to the browser immediately as a download attachment
    file_name = f"book_market_dashboard_{pages}_pages.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={file_name}"}
    )