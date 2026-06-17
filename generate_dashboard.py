import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

def create_visual_dashboard():
    desired_pages = 2  # Change this to pull more or less data dynamically!
    
    try:
        response = requests.get(f"http://127.0.0.1:8000/api/v1/books?pages={desired_pages}")
        api_data = response.json()
        books = api_data["data"]
    except Exception as e:
        print("Error connecting to live local API. Ensure Uvicorn is running!")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Overview"
    ws.views.sheetView[0].showGridLines = True

    # Title Block
    ws.merge_cells("A1:C2")
    title_cell = ws["A1"]
    title_cell.value = "Automated Book Market Analysis"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["Book Title", "Price (GBP)", "Status Assessment"]
    ws.append([]) 
    ws.append(headers)

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")

    # Data Population
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

    start_row = 5
    for idx, book in enumerate(books):
        current_row = start_row + idx
        clean_price = float(book["price"].replace("£", "").strip())
        
        ws.cell(row=current_row, column=1, value=book["title"])
        price_cell = ws.cell(row=current_row, column=2, value=clean_price)
        price_cell.number_format = '"£"#,##0.00'
        
        ws.cell(row=current_row, column=3, value=f'=IF(B{current_row}>40, "Premium Tier", "Budget Tier")')

        for col_idx in range(1, 4):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if col_idx == 2 else ("center" if col_idx == 3 else "left"))
            if current_row % 2 == 0:
                cell.fill = zebra_fill

    # Summary Accounting Row
    summary_row = start_row + len(books)
    ws.cell(row=summary_row, column=1, value="Average Market Price").font = Font(name="Arial", size=10, bold=True)
    avg_cell = ws.cell(row=summary_row, column=2, value=f"=AVERAGE(B5:B{summary_row-1})")
    avg_cell.font = Font(name="Arial", size=10, bold=True)
    avg_cell.number_format = '"£"#,##0.00'
    avg_cell.alignment = Alignment(horizontal="right")
    
    double_bottom = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    ws.cell(row=summary_row, column=1).border = double_bottom
    ws.cell(row=summary_row, column=2).border = double_bottom

    # Visual Chart Block
    chart = BarChart()
    chart.type = "col"
    chart.title = "Price Comparison per Book"
    chart.y_axis.title = "Price (£)"
    chart.x_axis.title = "Book Titles"
    chart.legend = None 
    chart.width = 25
    chart.height = 12

    data_ref = Reference(ws, min_col=2, min_row=4, max_row=summary_row-1)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=summary_row-1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    ws.add_chart(chart, "E4")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output_filename = "book_market_dashboard.xlsx"
    wb.save(output_filename)
    print(f"Success! Visual dashboard saved to '{output_filename}'")

if __name__ == "__main__":
    create_visual_dashboard()